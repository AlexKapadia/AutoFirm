"""Structural + determinism tests for the Excel financial-model builder.

These tests have teeth: every generated workbook is RE-OPENED with openpyxl and
its real structure asserted — sheet names, the exact live formulas wiring
calculations to driver cells, number formats, the FAST input fill, and the
period header — not merely that ``build`` did not raise. Determinism is proven by
byte-identical re-runs (CLAUDE.md §3.6); fail-closed paths are exercised.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st

from autofirm.artifacts.artifact_spec_validation_errors import ArtifactSpecError
from autofirm.artifacts.excel_financial_model_builder import (
    _column_letter,
    build_excel_financial_model,
)
from autofirm.artifacts.financial_model_spec import (
    CalculationRow,
    FinancialModelSpec,
    InputDriver,
)


def _model() -> FinancialModelSpec:
    return FinancialModelSpec(
        title="Acme",
        periods=("FY24", "FY25"),
        drivers=(
            InputDriver("revenue", "Revenue", (Decimal("100.00"), Decimal("120.00"))),
            InputDriver("cogs", "COGS", (Decimal("40.00"), Decimal("45.00")), "0.00"),
        ),
        calculations=(
            CalculationRow("gross_profit", "Gross profit", "={revenue} - {cogs}"),
            CalculationRow("margin", "Margin", "={gross_profit} / {revenue}", "0.0%"),
        ),
        outputs=("gross_profit", "margin"),
    )


def test_builder_creates_three_fast_sheets(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Inputs", "Calculations", "Outputs"]


def test_input_values_land_on_inputs_sheet(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    inputs = openpyxl.load_workbook(path)["Inputs"]
    # Row 3 = first driver (revenue), col B/C = FY24/FY25.
    assert inputs["A3"].value == "Revenue"
    assert inputs["B3"].value == 100
    assert inputs["C3"].value == 120
    assert inputs["A4"].value == "COGS"
    assert inputs["B4"].value == 40


def test_calculation_cells_are_live_formulas_referencing_inputs(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    calc = openpyxl.load_workbook(path)["Calculations"]
    # Gross profit row 3 must reference the Inputs cells for the SAME period,
    # never embed the numbers (FAST: no embedded constants).
    assert calc["B3"].value == "='Inputs'!B3 - 'Inputs'!B4"
    assert calc["C3"].value == "='Inputs'!C3 - 'Inputs'!C4"
    # Margin row 4 references the gross-profit calc cell + revenue input cell.
    assert calc["B4"].value == "='Calculations'!B3 / 'Inputs'!B3"


def test_outputs_reference_model_live(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    outputs = openpyxl.load_workbook(path)["Outputs"]
    assert outputs["A3"].value == "Gross profit"
    assert outputs["B3"].value == "='Calculations'!B3"
    assert outputs["B4"].value == "='Calculations'!B4"


def test_number_formats_applied(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb["Inputs"]["B4"].number_format == "0.00"  # COGS custom format
    assert wb["Calculations"]["B4"].number_format == "0.0%"  # margin custom format
    assert wb["Inputs"]["B3"].number_format == "#,##0.00"  # default revenue format


def test_period_header_row_present(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    inputs = openpyxl.load_workbook(path)["Inputs"]
    assert inputs["A2"].value == "Line item"
    assert inputs["B2"].value == "FY24"
    assert inputs["C2"].value == "FY25"


def test_input_cells_carry_fast_fill(tmp_path: Path) -> None:
    # FAST semantic colour coding: hard inputs are visually flagged.
    path = build_excel_financial_model(_model(), tmp_path / "m.xlsx")
    inputs = openpyxl.load_workbook(path)["Inputs"]
    assert inputs["B3"].fill.fgColor.rgb.endswith("FFF2CC")
    # Calculation cells must NOT carry the input fill (they are derived).
    calc = openpyxl.load_workbook(path)["Calculations"]
    assert not str(calc["B3"].fill.fgColor.rgb).endswith("FFF2CC")


def test_build_is_deterministic_byte_for_byte(tmp_path: Path) -> None:
    a = build_excel_financial_model(_model(), tmp_path / "a.xlsx")
    b = build_excel_financial_model(_model(), tmp_path / "b.xlsx")
    # openpyxl/zip metadata is timestamp-free for our content, so identical specs
    # must yield identical bytes (determinism — CLAUDE.md §3.6).
    assert a.read_bytes() == b.read_bytes()


@pytest.mark.parametrize("name", ["m.txt", "m.xls", "m", "m.XLSX.bak", "model.csv"])
def test_non_xlsx_destination_refused(tmp_path: Path, name: str) -> None:
    with pytest.raises(ArtifactSpecError, match=r"must be a \.xlsx"):
        build_excel_financial_model(_model(), tmp_path / name)


def test_uppercase_xlsx_extension_accepted(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "M.XLSX")
    assert path.exists()


def test_creates_missing_parent_directory(tmp_path: Path) -> None:
    path = build_excel_financial_model(_model(), tmp_path / "deep" / "nested" / "m.xlsx")
    assert path.exists()


@pytest.mark.parametrize(
    ("index", "expected"),
    [(1, "A"), (2, "B"), (26, "Z"), (27, "AA"), (28, "AB"), (52, "AZ"), (53, "BA"), (702, "ZZ")],
)
def test_column_letter_mapping(index: int, expected: str) -> None:
    assert _column_letter(index) == expected


@pytest.mark.property
@settings(max_examples=40, suppress_health_check=[HealthCheck.function_scoped_fixture])
@given(
    n_periods=st.integers(min_value=1, max_value=5),
    n_drivers=st.integers(min_value=1, max_value=4),
)
def test_every_driver_value_reaches_its_cell(
    tmp_path: Path, n_periods: int, n_drivers: int
) -> None:
    drivers = tuple(
        InputDriver(
            f"d{i}",
            f"Driver {i}",
            tuple(Decimal(i * 10 + p) for p in range(n_periods)),
        )
        for i in range(n_drivers)
    )
    spec = FinancialModelSpec(
        title="Gen", periods=tuple(f"P{p}" for p in range(n_periods)), drivers=drivers
    )
    path = build_excel_financial_model(spec, tmp_path / "g.xlsx")
    inputs = openpyxl.load_workbook(path)["Inputs"]
    # Every driver value must appear at its mapped cell — no off-by-one drift over
    # arbitrary model shapes (generality — CLAUDE.md §3.9).
    for di, driver in enumerate(drivers):
        for pi, value in enumerate(driver.values):
            cell = inputs.cell(row=3 + di, column=2 + pi)
            assert cell.value == int(value)
