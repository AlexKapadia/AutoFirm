"""Build a FAST-structured ``.xlsx`` financial model from a typed spec.

What this does
--------------
Renders a :class:`~autofirm.artifacts.financial_model_spec.FinancialModelSpec`
into a real Excel workbook with three sheets — **Inputs**, **Calculations**,
**Outputs** — laid out the FAST way (``docs/research/B15-artifact-generation``
§2.1): each driver lives once on Inputs; calculation cells are **live Excel
formulas** that reference the driver cells (never embedded constants); Outputs
references the model rather than re-deriving it. The result is a structured,
auditable model, not a dumped grid.

Why it exists / where it sits
-----------------------------
This is the L1.B15.1 Excel builder. It is intentionally library-thin: openpyxl
writes the formulas and formats; the *correctness* comes from the spec contract
(validated upstream) and the deterministic cell-mapping here. openpyxl writes
formula strings, not their computed results — a downstream LibreOffice-headless
recalc fills cached values; the model itself (formulas + structure) is exact and
deterministic given the spec.

Security / compliance invariants upheld
---------------------------------------
The builder validates the spec again at the boundary and refuses a non-``.xlsx``
target path *before* writing (fail-closed — CLAUDE.md §5.6). All labels are
written as text and all formulas are assembled only from validated keys, so no
spec field can inject an unintended formula. Output is byte-stable for identical
specs (determinism — CLAUDE.md §3.6).
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from autofirm.artifacts.artifact_spec_validation_errors import ArtifactSpecError
from autofirm.artifacts.financial_model_spec import _referenced_keys

if TYPE_CHECKING:
    from autofirm.artifacts.financial_model_spec import FinancialModelSpec

__all__ = ["build_excel_financial_model"]

# FAST semantic colour coding (§2.1): inputs are visually distinct from
# calculations so an auditor can see at a glance where every number originates.
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale amber = hard input
_HEADER_FILL = PatternFill("solid", fgColor="1F2A37")  # near-black header band
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LABEL_FONT = Font(bold=True)
_TITLE_FONT = Font(size=14, bold=True)

_FIRST_DATA_COL = 2  # column A holds labels; period values start at column B
_HEADER_ROW = 2  # row 1 is the sheet title; row 2 is the period header band
_FIRST_DATA_ROW = 3


def build_excel_financial_model(spec: FinancialModelSpec, destination: Path) -> Path:
    """Write ``spec`` to ``destination`` as a FAST-structured ``.xlsx`` file.

    Args:
        spec: A validated financial-model spec.
        destination: Target path; must end ``.xlsx``. Parent directories are
            created if missing.

    Returns:
        The ``destination`` path (now a written workbook).

    Raises:
        ArtifactSpecError: If ``destination`` is not a ``.xlsx`` path (fail-closed
            — CLAUDE.md §5.6). Spec-level invalidity is refused at spec
            construction.
    """
    if destination.suffix.lower() != ".xlsx":  # fail-closed: wrong container refused
        raise ArtifactSpecError(f"destination must be a .xlsx file, got {destination.name!r}")

    # Map every driver/calculation key to its row on its sheet so formulas can
    # reference the exact cell. Built once, used for all formula assembly.
    cell_index = _CellIndex(spec)

    workbook = Workbook()
    inputs_ws = workbook.active
    inputs_ws.title = "Inputs"
    calc_ws = workbook.create_sheet("Calculations")
    outputs_ws = workbook.create_sheet("Outputs")

    _write_inputs_sheet(inputs_ws, spec)
    _write_calculations_sheet(calc_ws, spec, cell_index)
    _write_outputs_sheet(outputs_ws, spec, cell_index)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


class _CellIndex:
    """Maps each driver/calculation key to its ``'Sheet'!$Col$Row`` cell pattern.

    Resolving references through a single index keeps the formula assembly
    deterministic and guarantees calculations point at the canonical cell for a
    key (calculate-once-then-reference — ICAEW P15).
    """

    def __init__(self, spec: FinancialModelSpec) -> None:
        self._sheet: dict[str, str] = {}
        self._row: dict[str, int] = {}
        for offset, driver in enumerate(spec.drivers):
            self._sheet[driver.key] = "Inputs"
            self._row[driver.key] = _FIRST_DATA_ROW + offset
        for offset, calc in enumerate(spec.calculations):
            self._sheet[calc.key] = "Calculations"
            self._row[calc.key] = _FIRST_DATA_ROW + offset

    def reference(self, key: str, period_index: int) -> str:
        """Return the absolute cross-sheet cell reference for ``key`` in a period."""
        col_letter = _column_letter(_FIRST_DATA_COL + period_index)
        return f"'{self._sheet[key]}'!{col_letter}{self._row[key]}"


def _write_inputs_sheet(ws: Worksheet, spec: FinancialModelSpec) -> None:
    """Write the Inputs sheet: title, period header, one amber row per driver."""
    _write_sheet_title(ws, f"{spec.title} — Inputs", len(spec.periods))
    _write_period_header(ws, spec)
    for offset, driver in enumerate(spec.drivers):
        row = _FIRST_DATA_ROW + offset
        _write_label(ws, row, driver.label)
        for period_index, value in enumerate(driver.values):
            cell = ws.cell(row=row, column=_FIRST_DATA_COL + period_index, value=value)
            cell.number_format = driver.number_format
            cell.fill = _INPUT_FILL  # FAST: hard inputs are visually flagged
    _set_layout(ws, spec)


def _write_calculations_sheet(
    ws: Worksheet, spec: FinancialModelSpec, cell_index: _CellIndex
) -> None:
    """Write the Calculations sheet: one live-formula row per calculation."""
    _write_sheet_title(ws, f"{spec.title} — Calculations", len(spec.periods))
    _write_period_header(ws, spec)
    for offset, calc in enumerate(spec.calculations):
        row = _FIRST_DATA_ROW + offset
        _write_label(ws, row, calc.label)
        refs = _referenced_keys(calc.formula_template)
        for period_index in range(len(spec.periods)):
            formula = _resolve_formula(calc.formula_template, refs, cell_index, period_index)
            cell = ws.cell(row=row, column=_FIRST_DATA_COL + period_index, value=formula)
            cell.number_format = calc.number_format
    _set_layout(ws, spec)


def _write_outputs_sheet(ws: Worksheet, spec: FinancialModelSpec, cell_index: _CellIndex) -> None:
    """Write the Outputs sheet: each output key referenced live from its source."""
    _write_sheet_title(ws, f"{spec.title} — Outputs", len(spec.periods))
    _write_period_header(ws, spec)
    label_of = _label_lookup(spec)
    for offset, key in enumerate(spec.outputs):
        row = _FIRST_DATA_ROW + offset
        _write_label(ws, row, label_of[key])
        for period_index in range(len(spec.periods)):
            ref = cell_index.reference(key, period_index)
            # Outputs reference the model live (no re-derivation) so the report
            # can never drift from the calculations it summarises.
            ws.cell(row=row, column=_FIRST_DATA_COL + period_index, value=f"={ref}")
    _set_layout(ws, spec)


def _resolve_formula(
    template: str, refs: tuple[str, ...], cell_index: _CellIndex, period_index: int
) -> str:
    """Substitute each ``{key}`` placeholder with its period cell reference."""
    resolved = template
    for ref in refs:
        resolved = resolved.replace(f"{{{ref}}}", cell_index.reference(ref, period_index))
    return resolved


def _write_sheet_title(ws: Worksheet, text: str, period_count: int) -> None:
    title_cell = ws.cell(row=1, column=1, value=text)
    title_cell.font = _TITLE_FONT
    if period_count > 0:
        # Merge the title across label + every period column for a clean banner.
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=_FIRST_DATA_COL + period_count - 1
        )


def _write_period_header(ws: Worksheet, spec: FinancialModelSpec) -> None:
    corner = ws.cell(row=_HEADER_ROW, column=1, value="Line item")
    corner.font = _HEADER_FONT
    corner.fill = _HEADER_FILL
    for period_index, period in enumerate(spec.periods):
        cell = ws.cell(row=_HEADER_ROW, column=_FIRST_DATA_COL + period_index, value=period)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="right")


def _write_label(ws: Worksheet, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _LABEL_FONT


def _set_layout(ws: Worksheet, spec: FinancialModelSpec) -> None:
    """Widen the label column and freeze the header band for readability."""
    ws.column_dimensions["A"].width = 34
    for period_index in range(len(spec.periods)):
        ws.column_dimensions[_column_letter(_FIRST_DATA_COL + period_index)].width = 14
    # Freeze below the header and right of the label column so headers stay
    # visible while scrolling a large model.
    ws.freeze_panes = ws.cell(row=_FIRST_DATA_ROW, column=_FIRST_DATA_COL)


def _label_lookup(spec: FinancialModelSpec) -> dict[str, str]:
    label_of: dict[str, str] = {d.key: d.label for d in spec.drivers}
    for calc in spec.calculations:
        label_of[calc.key] = calc.label
    return label_of


def _column_letter(column_index: int) -> str:
    """Convert a 1-based column index to its Excel letter (1->A, 27->AA)."""
    letters = ""
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters
